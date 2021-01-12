import rkivacc

import os
import tempfile
import email.utils

import requests
import openpyxl

class RKIReport:

    def __get_column(row, map, column):
        if not column in map:
            return None
        value = row[map[column]].value
        if isinstance(value, str):
            return RKIReport.__strip_stars(value)
        else:
            return value

    def __extract_row(row, map, state = True):
        data = {}
        
        if state:
            data["state"] = RKIReport.__get_column(row, map, "state")
        
        data.update({
                "vaccinations": RKIReport.__get_column(row, map, "vaccinations"),
                "delta": RKIReport.__get_column(row, map, "delta"),
                "vaccinations_per_capita": RKIReport.__get_column(row, map, "vaccinations_per_capita"),
                "reasons": {
                    "age": RKIReport.__get_column(row, map, "reasons_age"),
                    "profession": RKIReport.__get_column(row, map, "reasons_profession"),
                    "medical": RKIReport.__get_column(row, map, "reasons_medical"),
                    "nursing_home": RKIReport.__get_column(row, map, "reasons_nursing_home")
                }
        })
                    
        return data

    def __strip_stars(str):
        while str[-1] == "*":
            str = str[0:-1]
        return str

    def __init__(self, report_xlsx_path, modified = None):
        workbook = openpyxl.load_workbook(filename = report_xlsx_path, data_only = True)
        sheet = workbook.worksheets[rkivacc.WORKSHEET_INDEX]
        
        self._states = {}
        self._modified = modified
        
        table_rows = sheet.iter_rows(min_row=rkivacc.TABLE_FIRST_ROW - 1, 
                                     max_row=rkivacc.TABLE_FIRST_ROW + rkivacc.TABLE_LENGTH - 1,
                                     max_col=10)
                                     
        header_row = next(table_rows)
        
        known_rows = {
            "RS": None,
            "Bundesland": "state",
            "Impfungen kumulativ": "vaccinations",
            "Differenz zum Vortag": "delta",    
            "Impfungen pro 1.000 Einwohner": "vaccinations_per_capita",
            "Indikation nach Alter": "reasons_age",
            "Berufliche Indikation": "reasons_profession",
            "Medizinische Indikation": "reasons_medical",
            "Pflegeheim-bewohnerIn": "reasons_nursing_home",
        }

        map = {}

        for cellIndex in range(0, len(header_row)):
            value = header_row[cellIndex].value
            if value is None:
                continue
            value = RKIReport.__strip_stars(value)
            if not value in known_rows:
                print("Warning: Found unknown header item '{}'. This might be bad news.".format(value))
                continue
            key = known_rows[value]
            if key is None:
                continue
            map[key] = cellIndex
            
                                     
        total_row = next(sheet.iter_rows(min_row=rkivacc.TABLE_FIRST_ROW + rkivacc.TABLE_LENGTH,
                                         max_row=rkivacc.TABLE_FIRST_ROW + rkivacc.TABLE_LENGTH,
                                         max_col=10))
        
        for row in table_rows:
            self._states[row[map["state"]].value] = RKIReport.__extract_row(row, map)

        self._total = RKIReport.__extract_row(total_row, map, state = False)
    
    def states(self):
        return list(self._states.keys())
        
    def state(self, name):
        return self._states[name]
        
    def all_states(self):
        return self._states.values()
    
    def total(self):
        return self._total

    def modified(self):
        return self._modified
           
    def download(target_path):
        try:
            response = requests.get(rkivacc.URL)
        except RequestException:
            raise RuntimeError("Failed to obtain statistics from server")
        
        if not response.ok:
            raise RuntimeError("Unexpected status code from server: {}".format(response.status_code))
        
        with open(target_path, "wb") as file:
            file.write(response.content)
        return email.utils.parsedate_to_datetime(response.headers['Last-Modified'])

    def obtain():
        temp_dir = tempfile.TemporaryDirectory()
        temp_file = os.path.join(temp_dir.name, "rkiReport.xlsx")
        last_modified = RKIReport.download(temp_file)
        report = RKIReport(temp_file, last_modified)
        temp_dir.cleanup()
        return report
